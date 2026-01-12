"""
主机管理服务
提供主机连接测试、状态检查、批量操作等功能
基于Fabric SSH管理器
"""
import logging
from io import BytesIO
from typing import Any, Dict, List, Optional

from django.utils import timezone
from django.db import transaction
from django.contrib.contenttypes.models import ContentType
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError as DRFValidationError

from .models import Host, HostGroup
from .fabric_ssh_manager import fabric_ssh_manager, FabricSSHError
from .serializers import HostSerializer
from utils.audit_service import AuditLogService
import os

logger = logging.getLogger(__name__)


class HostService:
    """主机管理服务"""

    RAW_HEADER_ALIASES = {
        'name': ['name', '主机名', '主机名称'],
        'port': ['port', '端口', 'ssh_port', 'ssh端口'],
        'os_type': ['os_type', '操作系统', '系统', 'os'],
        'account': ['account', '账号', '服务器账号', 'account_name', '账号名称'],
        'description': ['description', '描述', '备注', 'remark'],
        'group_names': ['groups', 'group', '分组', '分组名称', 'group_names'],
        'owner': ['owner', '负责人'],
        'department': ['department', '部门'],
        'public_ip': ['public_ip', '外网ip', '公网ip'],
        'internal_ip': ['internal_ip', '内网ip', '内网'],
        'tags': ['tags', 'tag', '标签'],
        'service_role': ['service_role', '服务角色'],
        'remarks': ['remarks', 'notes', '备注'],
    }
    HEADER_ALIASES = {
        field: [alias.lower() for alias in aliases]
        for field, aliases in RAW_HEADER_ALIASES.items()
    }

    RAW_OS_ALIASES = {
        'linux': ['linux'],
        'windows': ['windows', 'win', 'winserver'],
        'aix': ['aix'],
        'solaris': ['solaris', 'sunos'],
    }
    OS_ALIASES = {
        field: [alias.lower() for alias in aliases]
        for field, aliases in RAW_OS_ALIASES.items()
    }

    SUPPORTED_OPTIONAL_FIELDS = [
        'description',
        'public_ip', 'internal_ip', 'tags',
        'owner', 'department',
        'service_role', 'remarks'
    ]

    MAX_DETAIL_ENTRIES = 200

    @staticmethod
    def _normalize_header(value) -> str:
        if value is None:
            return ''
        return str(value).strip().lower()

    @staticmethod
    def _clean_str(value) -> str:
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    @staticmethod
    def _parse_port(value) -> int:
        if value in (None, ''):
            return 22
        try:
            port = int(float(value))
        except (TypeError, ValueError):
            raise ValueError("端口必须为数字")
        if port <= 0 or port > 65535:
            raise ValueError("端口必须在 1-65535 之间")
        return port

    @classmethod
    def _normalize_os_type(cls, value):
        text = cls._normalize_header(value)
        if not text:
            return None
        for key, aliases in cls.OS_ALIASES.items():
            if text == key or text in aliases:
                return key
        return None


    @staticmethod
    def _split_group_names(value) -> List[str]:
        if not value:
            return []
        if not isinstance(value, str):
            value = str(value)
        separators = [',', '，', ';', '；', '|', '/', '、']
        normalized = value
        for sep in separators:
            normalized = normalized.replace(sep, ',')
        return [part.strip() for part in normalized.split(',') if part.strip()]

    @staticmethod
    def _split_tags(value) -> List[Dict[str, str]]:
        """
        支持以下格式：
          - list/tuple 中的对象 {key,value}
          - 逗号/空格分隔的 key=value 或 key
          - dict 将转为 [{key,value}]
        """
        if value is None:
            return []
        result: List[Dict[str, str]] = []
        if isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    key = str(item.get('key', '')).strip()
                    if not key:
                        continue
                    val = '' if item.get('value') is None else str(item.get('value')).strip()
                    result.append({'key': key, 'value': val})
                else:
                    text = str(item or '').strip()
                    if not text:
                        continue
                    if '=' in text:
                        key, *rest = text.split('=')
                        key = key.strip()
                        val = '='.join(rest).strip()
                    else:
                        key = text
                        val = ''
                    if key:
                        result.append({'key': key, 'value': val})
            return result

        if isinstance(value, dict):
            for k, v in value.items():
                key = str(k).strip()
                if not key:
                    continue
                result.append({'key': key, 'value': '' if v is None else str(v).strip()})
            return result

        raw_values: List[str] = [str(value)]
        separators = [',', '，', ';', '；', '|', '/', '、', ' ']
        for raw in raw_values:
            normalized = raw
            for sep in separators:
                normalized = normalized.replace(sep, ',')
            for part in normalized.split(','):
                part = part.strip()
                if not part:
                    continue
                if '=' in part:
                    key, *rest = part.split('=')
                    key = key.strip()
                    val = '='.join(rest).strip()
                else:
                    key = part
                    val = ''
                if key:
                    result.append({'key': key, 'value': val})
        return result

    @classmethod
    def _build_header_map(cls, headers) -> Dict[str, int]:
        header_map = {}
        for idx, header in enumerate(headers):
            normalized = cls._normalize_header(header)
            if not normalized:
                continue
            for field, aliases in cls.HEADER_ALIASES.items():
                if normalized == field or normalized in aliases:
                    if field not in header_map:
                        header_map[field] = idx
                    break
        return header_map

    @staticmethod
    def _format_validation_error(exc: DRFValidationError) -> str:
        if isinstance(exc.detail, dict):
            parts = []
            for field, messages in exc.detail.items():
                if isinstance(messages, list):
                    joined = ','.join(str(msg) for msg in messages)
                else:
                    joined = str(messages)
                parts.append(f"{field}: {joined}")
            return '; '.join(parts)
        if isinstance(exc.detail, list):
            return '; '.join(str(item) for item in exc.detail)
        return str(exc.detail)

    @staticmethod
    def _is_empty_row(row_values) -> bool:
        for value in row_values:
            if value is None:
                continue
            if isinstance(value, str) and value.strip() == '':
                continue
            return False
        return True

    @classmethod
    def _resolve_groups(cls, group_value, default_group: Optional[HostGroup], cache) -> (List[int], List[str]):
        names = cls._split_group_names(group_value)
        resolved_ids: List[int] = []
        missing_names: List[str] = []

        for name in names:
            key = name.lower()
            if key in cache:
                group = cache[key]
            else:
                group = HostGroup.objects.filter(name__iexact=name).first()
                cache[key] = group
            if group:
                if group.id not in resolved_ids:
                    resolved_ids.append(group.id)
            else:
                missing_names.append(name)

        if not resolved_ids and default_group:
            resolved_ids = [default_group.id]

        return resolved_ids, missing_names

    @classmethod
    def _build_payload(cls, *, row_data, name: str, internal_ip: Optional[str], 
                       public_ip: Optional[str], account_id: Optional[int],
                       port: int, existing: Optional[Host], default_group: Optional[HostGroup],
                       group_cache) -> (Dict[str, Any], List[str]):
        payload: Dict[str, Any] = {
            'name': name,
            'port': port,
        }
        
        # IP地址：至少需要一个
        if internal_ip:
            payload['internal_ip'] = internal_ip
        if public_ip:
            payload['public_ip'] = public_ip
        
        # 账号ID
        if account_id:
            payload['account'] = account_id

        os_value = cls._normalize_os_type(row_data.get('os_type'))
        if os_value:
            payload['os_type'] = os_value
        elif not existing:
            payload['os_type'] = 'linux'

        for field in cls.SUPPORTED_OPTIONAL_FIELDS:
            value = row_data.get(field)
            if value in (None, ''):
                continue

            if field == 'tags':
                tags = cls._split_tags(value)
                if tags:
                    payload[field] = tags
            else:
                payload[field] = cls._clean_str(value)

        group_ids, missing_groups = cls._resolve_groups(
            row_data.get('group_names'),
            default_group,
            group_cache
        )
        if group_ids:
            payload['groups'] = group_ids

        return payload, missing_groups
    
    @staticmethod
    def test_host_connection(host: Host, user=None) -> Dict[str, Any]:
        """测试主机连接，优先使用内网IP（短超时），失败后尝试外网IP"""
        use_internal = bool(host.internal_ip)
        use_public = bool(host.public_ip and host.public_ip != host.internal_ip)

        if not use_internal and not use_public:
            return {
                'success': False,
                'message': '主机没有配置IP地址（内网IP或外网IP）',
                'error_type': 'no_ip_address'
            }

        original_internal_ip = host.internal_ip
        original_public_ip = host.public_ip
        last_error = None
        used_ip_type = None
        used_ip = None

        # 优先尝试内网IP（短超时）
        if use_internal:
            try:
                host.internal_ip = original_internal_ip
                host.public_ip = None

                result = fabric_ssh_manager.execute_script(
                    host=host,
                    script_content='echo "connection_test"',
                    script_type='shell',
                    timeout=3,
                    connection_timeout=3
                )
            except (FabricSSHError, Exception) as e:
                last_error = str(e)
                used_ip_type = 'internal'
                used_ip = original_internal_ip
                logger.debug(f"内网IP连接失败: {host.name} ({original_internal_ip}) - {e}")
                result = None
            else:
                if result and result.get('success'):
                    host.status = 'online'
                    host.last_check_time = timezone.now()
                    host.save(update_fields=['status', 'last_check_time'])

                    host.internal_ip = original_internal_ip
                    host.public_ip = original_public_ip

                    if user:
                        AuditLogService.log_action(
                            user=user,
                            action='test_connection',
                            description=f'测试主机连接: {host.name} - 成功（使用内网IP）',
                            ip_address='127.0.0.1',
                            success=True,
                            resource_type=ContentType.objects.get_for_model(host),
                            resource_id=host.id,
                            resource_name=host.name
                        )

                    return {
                        'success': True,
                        'message': f'连接测试完成（使用内网IP: {original_internal_ip}）',
                        'stdout': result.get('stdout', ''),
                        'stderr': result.get('stderr', ''),
                        'used_ip': original_internal_ip,
                        'used_ip_type': 'internal',
                        'error_details': '',
                        'connection_info': result.get('connection_info', {}),
                    }
                else:
                    last_error = (result.get('message') if result else '内网连接失败')
                    used_ip_type = 'internal'
                    used_ip = original_internal_ip

        # 如果内网失败或没有内网IP，尝试外网IP
        if (not use_internal or last_error) and use_public:
            try:
                host.internal_ip = original_public_ip
                host.public_ip = original_public_ip

                result = fabric_ssh_manager.execute_script(
                    host=host,
                    script_content='echo "connection_test"',
                    script_type='shell',
                    timeout=10
                )
            except (FabricSSHError, Exception) as e:
                result = None
                last_error = str(e)
                logger.debug(f"外网IP连接失败: {host.name} ({original_public_ip}) - {e}")
            else:
                # 恢复原始IP设置
                host.internal_ip = original_internal_ip
                host.public_ip = original_public_ip

                if result and result.get('success'):
                    host.status = 'online'
                    host.last_check_time = timezone.now()
                    host.save(update_fields=['status', 'last_check_time'])

                    if user:
                        ip_info = "（使用外网IP）"
                        AuditLogService.log_action(
                            user=user,
                            action='test_connection',
                            description=f'测试主机连接: {host.name} - 成功{ip_info}',
                            ip_address='127.0.0.1',
                            success=True,
                            resource_type=ContentType.objects.get_for_model(host),
                            resource_id=host.id,
                            resource_name=host.name
                        )

                    return {
                        'success': True,
                        'message': f'连接测试完成（使用外网IP: {original_public_ip}）',
                        'stdout': result.get('stdout', ''),
                        'stderr': result.get('stderr', ''),
                        'used_ip': original_public_ip,
                        'used_ip_type': 'public',
                        'error_details': '',
                        'connection_info': result.get('connection_info', {}),
                    }
                else:
                    last_error = (result.get('message') if result else None) or (result.get('error') if result else None) or last_error or '外网连接失败'

        # 恢复原始IP地址
        host.internal_ip = original_internal_ip
        host.public_ip = original_public_ip

        # 所有IP都连接失败
        host.status = 'offline'
        host.save(update_fields=['status'])

        result = {
            'success': False,
            'message': f'连接失败: {last_error}',
            'error_type': 'connection_error',
            'error_details': last_error or '',
            'connection_info': {
                'used_ip': used_ip,
                'used_ip_type': used_ip_type,
            }
        }

        if user:
            AuditLogService.log_action(
                user=user,
                action='test_connection',
                description=f'测试主机连接失败: {host.name} - {last_error}',
                ip_address='127.0.0.1',
                success=False,
                resource_type=ContentType.objects.get_for_model(host),
                resource_id=host.id,
                resource_name=host.name
            )

        return result
    
    @staticmethod
    def batch_test_connections(hosts: List[Host], user=None) -> Dict[str, Any]:
        """批量测试主机连接"""
        results = {
            'total': len(hosts),
            'success': 0,
            'failed': 0,
            'details': []
        }
        
        for host in hosts:
            result = HostService.test_host_connection(host, user)
            
            if result['success']:
                results['success'] += 1
            else:
                results['failed'] += 1
            
            results['details'].append({
                'host_id': host.id,
                'host_name': host.name,
                'ip_address': host.ip_address,
                'result': result
            })
        
        return results
    
    @staticmethod
    def check_host_status(host: Host) -> str:
        """检查主机状态"""
        try:
            result = fabric_ssh_manager.execute_script(
                host=host,
                script_content='echo "status_check"',
                script_type='shell',
                timeout=10,
                connection_timeout=5
            )
            return 'online' if result['success'] else 'offline'
        except Exception:
            return 'offline'
    
    @staticmethod
    def batch_check_status(hosts: List[Host]) -> Dict[str, int]:
        """批量检查主机状态"""
        status_count = {
            'online': 0,
            'offline': 0,
            'unknown': 0
        }
        
        for host in hosts:
            status = HostService.check_host_status(host)
            host.status = status
            host.last_check_time = timezone.now()
            host.save(update_fields=['status', 'last_check_time'])
            
            status_count[status] += 1
        
        return status_count
    
    @staticmethod
    def get_host_system_info(host: Host) -> Dict[str, Any]:
        """获取主机系统信息"""
        try:
            # 使用Fabric获取系统信息
            script_content = """
echo "=== OS Release ==="
cat /etc/os-release 2>/dev/null || echo "N/A"
echo "=== CPU Info ==="
lscpu 2>/dev/null || echo "N/A"
echo "=== Memory Info ==="
free -h 2>/dev/null || echo "N/A"
echo "=== Disk Info ==="
df -h 2>/dev/null || echo "N/A"
"""
            
            result = fabric_ssh_manager.execute_script(
                host=host,
                script_content=script_content,
                script_type='shell',
                timeout=30,
                connection_timeout=10
            )

            if result['success']:
                return {
                    'success': True,
                    'system_info': {
                        'raw_output': result['stdout']
                    }
                }
            else:
                return {
                    'success': False,
                    'message': result.get('message', '获取系统信息失败'),
                    'system_info': {}
                }
                
        except Exception as e:
            logger.error(f"获取主机系统信息失败: {host.name} - {e}")
            return {
                'success': False,
                'message': f'获取系统信息失败: {e}',
                'system_info': {}
            }
    
    @staticmethod
    def collect_system_info(host: Host, user=None) -> Dict[str, Any]:
        """收集主机系统信息，优先使用内网IP（短超时），失败后尝试外网IP"""
        # 确定使用的IP地址（优先内网IP）
        use_internal = bool(host.internal_ip)
        use_public = bool(host.public_ip and host.public_ip != host.internal_ip)
        
        if not use_internal and not use_public:
            return {
                'success': False,
                'message': '主机没有配置IP地址（内网IP或外网IP）',
                'system_info': {}
            }
        
        original_internal_ip = host.internal_ip
        original_public_ip = host.public_ip
        # 从外部脚本加载收集脚本
        script_path = os.path.join(os.path.dirname(__file__), "scripts", "collect_system_info.sh")
        try:
            with open(script_path, "r", encoding="utf-8") as fh:
                collect_script = fh.read()
        except Exception as e:
            logger.exception("加载收集系统信息脚本失败: %s", e)
            return {
                'success': False,
                'message': f'加载收集脚本失败: {e}',
                'system_info': {}
            }

        result = None
        
        # 优先尝试内网IP（使用短超时）
        if use_internal:
            try:
                host.internal_ip = original_internal_ip
                host.public_ip = None  # 避免使用公网IP

                # 内网连接使用短超时（连接5秒，命令执行5秒，收集信息需要执行脚本，比测试连接稍长）
                result = fabric_ssh_manager.execute_script(
                    host=host,
                    script_content=collect_script,
                    script_type='shell',
                    timeout=5,
                    connection_timeout=5  # 连接超时也设置为5秒
                )
            except (FabricSSHError, Exception) as e:
                last_error = str(e)
                result = None
                logger.debug(f"内网IP收集信息异常: {host.name} ({original_internal_ip}) - {e}")
            else:
                if not result or not result.get('success'):
                    last_error = result.get('message') if result else '内网连接失败'
                    result = None
                    logger.debug(f"内网IP收集信息失败: {host.name} ({original_internal_ip}) - {last_error}")
        
        # 如果内网失败或没有内网IP，尝试外网IP
        if (not use_internal or not result) and use_public:
            try:
                # 临时使用公网IP作为连接目标
                host.internal_ip = original_public_ip
                host.public_ip = original_public_ip

                # 外网连接使用正常超时（60秒）
                result = fabric_ssh_manager.execute_script(
                    host=host,
                    script_content=collect_script,
                    script_type='shell',
                    timeout=60
                )
            except (FabricSSHError, Exception) as e:
                last_error = str(e)
                result = None
                logger.debug(f"外网IP收集信息异常: {host.name} ({original_public_ip}) - {e}")
            else:
                if not result or not result.get('success'):
                    last_error = result.get('message') if result else '外网连接失败'
                    logger.debug(f"外网IP收集信息失败: {host.name} ({original_public_ip}) - {last_error}")
        
        # 恢复原始IP地址
        host.internal_ip = original_internal_ip
        host.public_ip = original_public_ip

        # 检查是否成功
        if not result or not result.get('success'):
            error_msg = last_error or '未知错误'
            return {
                'success': False,
                'message': f'系统信息收集失败: {error_msg}',
                'system_info': {}
            }

        # 解析输出
        try:
            output = result.get('stdout', '') or result.get('output', '')
            system_info = {}

            # 提取系统信息
            lines = output.split('\n')
            in_info_section = False

            for line in lines:
                line = line.strip()
                if line == "=== SYSTEM_INFO_START ===":
                    in_info_section = True
                    continue
                elif line == "=== SYSTEM_INFO_END ===":
                    break
                elif in_info_section and '=' in line:
                    key, value = line.split('=', 1)
                    if value and value != 'unknown':
                        system_info[key] = value

            # 更新主机信息
            update_fields = []

            if 'HOSTNAME' in system_info:
                host.hostname = system_info['HOSTNAME']
                update_fields.append('hostname')

            if 'OS_VERSION' in system_info:
                host.os_version = system_info['OS_VERSION']
                update_fields.append('os_version')

            if 'KERNEL_VERSION' in system_info:
                host.kernel_version = system_info['KERNEL_VERSION']
                update_fields.append('kernel_version')

            if 'CPU_CORES' in system_info and system_info['CPU_CORES'].isdigit():
                host.cpu_cores = int(system_info['CPU_CORES'])
                update_fields.append('cpu_cores')

            if 'MEMORY_GB' in system_info and system_info['MEMORY_GB'].isdigit():
                host.memory_gb = float(system_info['MEMORY_GB'])
                update_fields.append('memory_gb')

            if 'DISK_GB' in system_info and system_info['DISK_GB'].isdigit():
                host.disk_gb = float(system_info['DISK_GB'])
                update_fields.append('disk_gb')

            if 'INTERNAL_IP' in system_info:
                host.internal_ip = system_info['INTERNAL_IP']
                update_fields.append('internal_ip')

            if 'CLOUD_PROVIDER' in system_info and system_info['CLOUD_PROVIDER'] != 'unknown':
                host.cloud_provider = system_info['CLOUD_PROVIDER']
                update_fields.append('cloud_provider')

            if 'INSTANCE_ID' in system_info:
                host.instance_id = system_info['INSTANCE_ID']
                update_fields.append('instance_id')

            if 'REGION' in system_info:
                host.region = system_info['REGION']
                update_fields.append('region')

            if 'ZONE' in system_info:
                host.zone = system_info['ZONE']
                update_fields.append('zone')

            # 网络信息
            if 'PUBLIC_IP' in system_info:
                host.public_ip = system_info['PUBLIC_IP']
                update_fields.append('public_ip')

            if 'INTERNAL_MAC' in system_info:
                host.internal_mac = system_info['INTERNAL_MAC']
                update_fields.append('internal_mac')

            if 'EXTERNAL_MAC' in system_info:
                host.external_mac = system_info['EXTERNAL_MAC']
                update_fields.append('external_mac')

            if 'GATEWAY' in system_info:
                host.gateway = system_info['GATEWAY']
                update_fields.append('gateway')

            if 'DNS_SERVERS' in system_info:
                host.dns_servers = system_info['DNS_SERVERS']
                update_fields.append('dns_servers')

            # 系统详细信息
            if 'CPU_MODEL' in system_info:
                host.cpu_model = system_info['CPU_MODEL'][:200]  # 限制长度
                update_fields.append('cpu_model')

            if 'OS_ARCH' in system_info:
                host.os_arch = system_info['OS_ARCH']
                update_fields.append('os_arch')

            if 'CPU_ARCH' in system_info:
                host.cpu_arch = system_info['CPU_ARCH']
                update_fields.append('cpu_arch')

            # 更新最后检查时间
            host.last_check_time = timezone.now()
            update_fields.append('last_check_time')

            # 保存更新
            if update_fields:
                host.save(update_fields=update_fields)

            # 记录操作日志
            if user:
                AuditLogService.log_action(
                    user=user,
                    action='collect_system_info',
                    description=f'收集主机系统信息: {host.name} - 成功更新 {len(update_fields)} 个字段',
                    ip_address='127.0.0.1',
                    success=True,
                    resource_type=ContentType.objects.get_for_model(host),
                    resource_id=host.id,
                    resource_name=host.name,
                    extra_data={'updated_fields': update_fields, 'system_info': system_info}
                )

            return {
                'success': True,
                'message': f'系统信息收集成功，更新了 {len(update_fields)} 个字段',
                'system_info': system_info,
                'updated_fields': update_fields
            }

        except Exception as e:
            logger.error(f"收集主机系统信息失败: {host.name} - {e}")
            return {
                'success': False,
                'message': f'收集系统信息失败: {e}',
                'system_info': {}
            }

    @staticmethod
    def execute_command_on_host(host: Host, command: str, timeout: int = 30, user=None) -> Dict[str, Any]:
        """在主机上执行命令"""
        try:
            result = fabric_ssh_manager.execute_script(
                host=host,
                script_content=command,
                script_type='shell',
                timeout=timeout,
                connection_timeout=10
            )

            # 记录操作日志
            if user:
                AuditLogService.log_action(
                    user=user,
                    action='execute_script',
                    description=f'执行命令: {command} (退出码: {result["exit_code"]})',
                    ip_address='127.0.0.1',
                    success=result['success'],
                    resource_type=ContentType.objects.get_for_model(host),
                    resource_id=host.id,
                    resource_name=host.name,
                    extra_data={'command': command, 'exit_code': result['exit_code']}
                )
            
            return {
                'success': result['success'],
                'exit_code': result['exit_code'],
                'stdout': result['stdout'],
                'stderr': result['stderr'],
                'command': command
            }
            
        except Exception as e:
            logger.error(f"执行命令失败: {host.name} - {e}")
            return {
                'success': False,
                'error': str(e),
                'command': command
            }

    @classmethod
    def import_hosts_from_excel(cls, uploaded_file, user, default_group: Optional[HostGroup] = None,
                                overwrite_existing: bool = False) -> Dict[str, Any]:
        """通过Excel批量导入主机"""
        if not uploaded_file:
            return {'success': False, 'message': '未上传Excel文件'}

        try:
            file_bytes = uploaded_file.read()
            buffer = BytesIO(file_bytes)
            workbook = load_workbook(filename=buffer, data_only=True)
        except Exception as exc:
            logger.exception("解析Excel失败: %s", exc)
            return {'success': False, 'message': f'解析Excel失败: {exc}'}
        finally:
            if hasattr(uploaded_file, 'seek'):
                try:
                    uploaded_file.seek(0)
                except Exception:
                    pass

        summary = {
            'total': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'failed': 0,
        }
        row_results: List[Dict[str, Any]] = []
        group_cache: Dict[str, Optional[HostGroup]] = {}

        try:
            sheet = workbook.active
            header_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            try:
                header_row = next(header_iter)
            except StopIteration:
                return {'success': False, 'message': 'Excel文件缺少表头'}

            header_map = cls._build_header_map(header_row)
            required_fields = ['name', 'account']
            missing_columns = [field for field in required_fields if field not in header_map]
            if missing_columns:
                return {
                    'success': False,
                    'message': f"Excel缺少必要列: {', '.join(missing_columns)}",
                    'missing_columns': missing_columns
                }
            
            # 检查至少有一个IP地址字段
            has_ip = 'internal_ip' in header_map or 'public_ip' in header_map
            if not has_ip:
                return {
                    'success': False,
                    'message': "Excel缺少IP地址列（至少需要 internal_ip 或 public_ip 之一）",
                    'missing_columns': ['internal_ip', 'public_ip']
                }

            # 账号缓存
            from .models import ServerAccount
            account_cache: Dict[str, Optional[ServerAccount]] = {}

            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if cls._is_empty_row(row):
                        continue

                    summary['total'] += 1
                    row_data = {field: row[col_idx] for field, col_idx in header_map.items()}

                    name = cls._clean_str(row_data.get('name'))
                    internal_ip = cls._clean_str(row_data.get('internal_ip'))
                    public_ip = cls._clean_str(row_data.get('public_ip'))
                    account_name = cls._clean_str(row_data.get('account'))

                    row_result = {
                        'row': row_idx,
                        'name': name or None,
                        'internal_ip': internal_ip or None,
                        'public_ip': public_ip or None,
                    }

                    try:
                        if not name:
                            raise ValueError('主机名称不能为空')
                        if not internal_ip and not public_ip:
                            raise ValueError('至少需要配置内网IP或外网IP之一')
                        if not account_name:
                            raise ValueError('服务器账号不能为空')
                        
                        # 查找账号
                        account_key = account_name.lower()
                        if account_key in account_cache:
                            account = account_cache[account_key]
                        else:
                            account = ServerAccount.objects.filter(name__iexact=account_name).first()
                            account_cache[account_key] = account
                        
                        if not account:
                            raise ValueError(f'服务器账号 "{account_name}" 不存在，请先在服务器账号管理中创建')
                        
                        port = cls._parse_port(row_data.get('port'))
                        # 查找现有主机：使用内网IP或外网IP+端口
                        existing = None
                        if internal_ip:
                            existing = Host.objects.filter(internal_ip=internal_ip, port=port).first()
                        if not existing and public_ip:
                            existing = Host.objects.filter(public_ip=public_ip, port=port).first()

                        if existing and not overwrite_existing:
                            summary['skipped'] += 1
                            row_result['status'] = 'skipped'
                            row_result['message'] = '主机已存在（根据 IP + 端口）'
                            row_results.append(row_result)
                            continue

                        payload, missing_groups = cls._build_payload(
                            row_data=row_data,
                            name=name,
                            internal_ip=internal_ip,
                            public_ip=public_ip,
                            account_id=account.id,
                            port=port,
                            existing=existing,
                            default_group=default_group,
                            group_cache=group_cache,
                        )

                        if existing:
                            serializer = HostSerializer(existing, data=payload, partial=True)
                        else:
                            serializer = HostSerializer(data=payload)

                        serializer.is_valid(raise_exception=True)
                        if existing:
                            serializer.save()
                            summary['updated'] += 1
                            row_result['status'] = 'updated'
                            row_result['message'] = '已更新现有主机'
                        else:
                            serializer.save(created_by=user)
                            summary['created'] += 1
                            row_result['status'] = 'created'
                            row_result['message'] = '创建成功'

                        if missing_groups:
                            row_result['missing_groups'] = missing_groups

                    except DRFValidationError as exc:
                        summary['failed'] += 1
                        row_result['status'] = 'failed'
                        row_result['message'] = cls._format_validation_error(exc)
                    except ValueError as exc:
                        summary['failed'] += 1
                        row_result['status'] = 'failed'
                        row_result['message'] = str(exc)
                    except Exception as exc:
                        summary['failed'] += 1
                        row_result['status'] = 'failed'
                        row_result['message'] = str(exc)
                        logger.exception("导入主机失败 (行 %s): %s", row_idx, exc)

                    row_results.append(row_result)
        finally:
            workbook.close()

        details = row_results[:cls.MAX_DETAIL_ENTRIES]
        message = (
            f"导入完成：新增 {summary['created']} 条，"
            f"更新 {summary['updated']} 条，跳过 {summary['skipped']} 条，"
            f"失败 {summary['failed']} 条"
        )

        result = {
            'summary': summary,
            'details': details,
            'message': message,
        }
        if len(row_results) > cls.MAX_DETAIL_ENTRIES:
            result['limit_note'] = (
                f"仅展示前 {cls.MAX_DETAIL_ENTRIES} 条明细，"
                f"剩余 {len(row_results) - cls.MAX_DETAIL_ENTRIES} 条请查看日志或源文件"
            )
        result['success'] = summary['failed'] == 0
        return result

    @classmethod
    def generate_excel_template(cls) -> bytes:
        """生成主机导入模板Excel"""
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hosts"

        headers = [
            "主机名称",
            "内网IP",
            "外网IP(可选)",
            "端口(可选, 默认22)",
            "操作系统(可选: linux/windows/aix/solaris)",
            "服务器账号",
            "分组(支持多个, 逗号分隔)",
            "描述",
        ]
        sheet.append(headers)

        sample_row = [
            "demo-host-01",
            "192.168.1.10",
            "10.0.0.10",
            22,
            "linux",
            "root账号",
            "默认分组,数据库",
            "演示主机",
        ]
        sheet.append(sample_row)

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)
        return buffer.read()


class HostGroupService:
    """主机分组管理服务"""
    
    @staticmethod
    def create_group(name: str, description: str = '', parent: HostGroup = None, 
                    sort_order: int = 0, user=None) -> HostGroup:
        """创建主机分组"""
        with transaction.atomic():
            group = HostGroup.objects.create(
                name=name,
                description=description,
                parent=parent,
                sort_order=sort_order,
                created_by=user
            )
            
            # 记录操作日志
            if user:
                AuditLogService.log_action(
                    user=user,
                    action='create',
                    description=f'创建主机分组: {name}',
                    ip_address='127.0.0.1',
                    resource_type=ContentType.objects.get_for_model(group),
                    resource_id=group.id,
                    resource_name=group.name
                )
            
            return group
